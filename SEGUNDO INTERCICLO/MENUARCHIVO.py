
import os
from openpyxl import Workbook, load_workbook

def crear_archivo(nombre, tipo):
    if tipo == 'txt':
        open(nombre, 'w').close()
    elif tipo == 'xlsx':
        wb = Workbook()
        wb.save(nombre)
    print(f"Archivo '{nombre}' ({tipo}) creado.")

def escribir_archivo(nombre, tipo):
    if not os.path.exists(nombre):
        print(f"El archivo '{nombre}' no existe.")
        return

    texto = input("Ingrese el texto a escribir: ")

    if tipo == 'txt':
        with open(nombre, 'a') as f:
            f.write(texto + '\n')
    elif tipo == 'xlsx':
        wb = load_workbook(nombre)
        ws = wb.active
        ws.append([texto])
        wb.save(nombre)

    print(f"Texto escrito en '{nombre}'.")

def eliminar_archivo(nombre):
    if os.path.exists(nombre):
        os.remove(nombre)
        print(f"Archivo '{nombre}' eliminado.")
    else:
        print(f"El archivo '{nombre}' no existe.")

def menu():
    archivo1 = "archivo1.txt"
    tipo1 = 'txt'
    archivo2 = None
    tipo2 = None

    while True:
        print("\n--- MENÚ ---")
        print("1. Crear archivo por defecto ('archivo1.txt')")
        print("2. Crear archivo nuevo (nombre y tipo personalizado)")
        print("3. Escribir en archivo por defecto")
        print("4. Escribir en archivo nuevo")
        print("5. Eliminar archivo por defecto")
        print("6. Eliminar archivo nuevo")
        print("7. Salir")
        opcion = input("Seleccione una opción: ")

        if opcion == '1':
            crear_archivo(archivo1, tipo1)
        elif opcion == '2':
            nombre = input("Ingrese el nombre del nuevo archivo (sin extensión): ")
            tipo = input("Seleccione tipo ('txt' o 'xlsx'): ").lower()
            if tipo not in ['txt', 'xlsx']:
                print("Tipo no válido.")
                continue
            archivo2 = f"{nombre}.{tipo}"
            tipo2 = tipo
            crear_archivo(archivo2, tipo2)
        elif opcion == '3':
            escribir_archivo(archivo1, tipo1)
        elif opcion == '4':
            if archivo2:
                escribir_archivo(archivo2, tipo2)
            else:
                print("Primero debe crear el archivo nuevo.")
        elif opcion == '5':
            eliminar_archivo(archivo1)
        elif opcion == '6':
            if archivo2:
                eliminar_archivo(archivo2)
            else:
                print("No hay archivo nuevo para eliminar.")
        elif opcion == '7':
            print("Saliendo del programa.")
            break
        else:
            print("Opción no válida.")

if __name__ == "__main__":
    print("Bienvenido al gestor de archivos (TXT/Excel).")
    menu()
